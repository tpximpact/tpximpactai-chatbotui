import logging
from typing import Iterator, List, Union
from pathlib import Path
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from langchain_core.documents import Document
from langchain_community.document_loaders.base import BaseLoader

class CustomExcelLoader(BaseLoader):
    """Load an Excel (.xlsx) file into a list of Documents using lightweight openpyxl.
    
    Processes all sheets in the workbook and formats content in a way that preserves context
    for AI models to understand the table structure, with clear sheet names and row numbers.
    """
    
    def __init__(self, file_path: Union[str, Path]):
        """Initialize the Excel loader.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = Path(file_path)
    
    def lazy_load(self) -> Iterator[Document]:
        """Load documents from Excel file in a lazy (streaming) fashion."""
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(
                self.file_path, read_only=True, data_only=True
            )
            
            all_sheets_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.rows)
                
                if not rows:
                    logging.info(f"Sheet {sheet_name} has no rows")
                    continue
                
                # Add sheet header
                sheet_content = [f"SHEET: \"{sheet_name}\""]
                
                # Process each row with its Excel row number
                for row_idx, row in enumerate(rows, 1):  # Excel is 1-indexed
                    row_content = [f"ROW {row_idx}:"]
                    
                    # Process each cell in the row
                    for col_idx, cell in enumerate(row, 1):  # Make columns 1-indexed too
                        value = cell.value
                        # Handle None values
                        if value is None:
                            value = ""
                        # Convert numeric and other types to string
                        if not isinstance(value, str):
                            value = str(value)
                            
                        # Add cell content with column letter (A, B, C, etc.)
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        row_content.append(f"CELL {col_letter}{row_idx}: {value}")
                    
                    # Join the row content with newlines and add to sheet content
                    sheet_content.append("\n".join(row_content))
                
                # Join all rows for this sheet with double newlines
                formatted_sheet = "\n\n".join(sheet_content)
                all_sheets_content.append(formatted_sheet)
            
            # Create a single document with all sheets' content
            yield Document(
                page_content="\n\n".join(all_sheets_content),
                metadata={
                    "source": str(self.file_path),
                }
            )
            
            # Close the workbook to release resources
            workbook.close()
                
        except InvalidFileException:
            raise ValueError(f"Not a valid Excel file: {self.file_path}")
        except Exception as e:
            raise RuntimeError(f"Error loading Excel file {self.file_path}: {e}")
    
    def load(self) -> List[Document]:
        """Load all documents from the Excel file."""
        return list(self.lazy_load())